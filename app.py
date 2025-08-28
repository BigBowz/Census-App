import os
import re
from io import BytesIO
from datetime import datetime

import pandas as pd
import streamlit as st

from amplify_census_input import normalize_and_validate
from export_aetna_afa import export_to_aetna_afa

# --------------------------------
# Page config FIRST (important)
# --------------------------------
st.set_page_config(page_title="Census Converter", layout="wide")

# --------------------------------
# Simple password gate (default Amplify#1; override via secret/env)
# --------------------------------
def check_password() -> bool:
    """
    Single-password gate.
    Default is 'Amplify#1', but you can override with:
      - Streamlit secret: app_password = "..."
      - Env var:          APP_PASSWORD=...
    """
    expected = (
        st.secrets.get("app_password")
        or os.environ.get("APP_PASSWORD")
        or "Amplify#1"
    )

    if st.session_state.get("auth_ok"):
        return True

    st.title("Census Converter 🔐")
    pw = st.text_input("Enter password", type="password")
    if st.button("Enter"):
        st.session_state["auth_ok"] = (pw == expected)
        if not st.session_state["auth_ok"]:
            st.error("Incorrect password")
            st.stop()
    else:
        st.stop()
    return True

# Call gate BEFORE rendering the rest of the app
check_password()

st.title("Census Converter")
st.caption("Upload → Pick Sheet & Rows → Normalize & Validate → Convert → Download")

# --------------------------------
# Helpers (with caching)
# --------------------------------
def _clean_str(v) -> str:
    s = "" if v is None else str(v).strip()
    return re.sub(r"\s+", " ", s)

@st.cache_data(show_spinner=False)
def get_sheet_names(file_bytes: bytes):
    """Return list of sheet names from an Excel file (cached)."""
    return pd.ExcelFile(BytesIO(file_bytes)).sheet_names

@st.cache_data(show_spinner=False)
def read_raw_preview(file_bytes: bytes, sheet: str, nrows: int = 30) -> pd.DataFrame:
    """Read first N rows without headers for visual selection (cached)."""
    return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None, nrows=nrows)

@st.cache_data(show_spinner=False)
def parse_with_rows(file_bytes: bytes, sheet: str, header_row_excel: int, first_data_row_excel: int) -> pd.DataFrame:
    """Parse using chosen header row & first data row (cached)."""
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=int(header_row_excel - 1))
    start_offset = int(first_data_row_excel - (header_row_excel + 1))
    if start_offset > 0:
        df = df.iloc[start_offset:].reset_index(drop=True)
    return df

def read_sheet_fields(file_bytes: bytes, sheet_name: str) -> dict:
    """
    Read company/address/fein/sic from the uploaded workbook (lazy import).
      C5 -> company
      C6 -> address
      C8 -> fein
      E7 -> sic
    """
    from openpyxl import load_workbook  # lazy import to speed cold starts
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        company = ws["C5"].value
        address = ws["C6"].value
        fein    = ws["C8"].value
        sic     = ws["E7"].value
    finally:
        wb.close()
    return {
        "company": _clean_str(company),
        "address": _clean_str(address),
        "fein":    _clean_str(fein),
        "sic":     _clean_str(sic),
    }

@st.cache_data(show_spinner=False)
def read_meta_fields_cached(file_bytes: bytes, sheet: str) -> dict:
    """Cached wrapper for header cells (C5/C6/C8/E7)."""
    return read_sheet_fields(file_bytes, sheet)

@st.cache_data(show_spinner=False)
def normalize_cached(df: pd.DataFrame):
    """Cached normalization + validation."""
    return normalize_and_validate(df)

@st.cache_data(show_spinner=False)
def convert_cached(norm_df: pd.DataFrame, meta: dict) -> bytes:
    """Cached conversion to Aetna AFA bytes."""
    return export_to_aetna_afa(norm_df, meta=meta)

def sanitize_filename_component(s: str) -> str:
    s = re.sub(r'[\\/:*?"<>|]+', "", s or "")
    s = re.sub(r"\s+", " ", s).strip()
    return s or "Unknown Company"

# --------------------------------
# App body
# --------------------------------
uploaded = st.file_uploader("Upload census Excel (.xlsx)", type=["xlsx"])

if uploaded is not None:
    file_bytes = uploaded.getvalue()

    # 1) Sheet picker
    try:
        sheets = get_sheet_names(file_bytes)
    except Exception as e:
        st.error(f"Could not open workbook: {e}")
        st.stop()

    with st.expander("Step 1 — Choose sheet & rows", expanded=True):
        colA, colB, colC = st.columns([1.2, 1, 1])

        with colA:
            sheet = st.selectbox("Sheet", options=sheets, index=0)

        # Raw preview (no header) to help pick rows
        try:
            raw_preview = read_raw_preview(file_bytes, sheet)
        except Exception as e:
            st.error(f"Could not read preview: {e}")
            st.stop()

        raw_preview_display = raw_preview.copy()
        raw_preview_display.index = raw_preview_display.index + 1  # Excel-like row numbers
        st.write("**Raw preview (first 30 rows, no headers yet)**")
        st.dataframe(raw_preview_display, use_container_width=True)

        max_row = max(30, len(raw_preview_display))
        with colB:
            header_row_excel = st.number_input(
                "Header row (Excel #)",
                min_value=1, max_value=max_row, value=1, step=1,
                help="Row that contains the column names",
            )
        with colC:
            first_data_row_excel = st.number_input(
                "First data row (Excel #)",
                min_value=header_row_excel + 1, max_value=max_row + 10,
                value=header_row_excel + 1, step=1,
                help="Usually header + 1",
            )

        # Parse with selections
        try:
            parsed_df = parse_with_rows(file_bytes, sheet, int(header_row_excel), int(first_data_row_excel))
        except Exception as e:
            st.error(f"Could not parse with selected rows: {e}")
            st.stop()

        st.write("**Parsed preview with your header/data selection (first 20 rows)**")
        st.dataframe(parsed_df.head(20), use_container_width=True)

    # 2) Read meta fields (for header cells + filename)
    meta = read_meta_fields_cached(file_bytes, sheet)
    company_for_name = sanitize_filename_component(meta.get("company", ""))
    date_str = datetime.now().date().isoformat()
    out_filename = f"Amplify AFA Census for {company_for_name} {date_str}.xlsx"

    # 3) Normalize & validate
    try:
        norm_df, issues_df = normalize_cached(parsed_df)
    except Exception as e:
        st.error(f"Normalization error: {e}")
        st.stop()

    st.subheader("Normalized input (first 20 rows)")
    st.dataframe(norm_df.head(20), use_container_width=True)

    if not issues_df.empty:
        st.warning(f"Validation found {len(issues_df)} issue(s). You can still convert, but review below:")
        st.dataframe(issues_df, use_container_width=True)
        issues_csv = issues_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "Download issues (CSV)",
            data=issues_csv,
            file_name="census_issues.csv",
            mime="text/csv",
        )
    else:
        st.success("No validation issues found.")

    # 4) Convert → Aetna AFA
    with st.spinner("Converting…"):
        out_bytes = convert_cached(norm_df, meta)

    # 5) Output preview (headers are on row 7 → header=6)
    try:
        out_preview = pd.read_excel(BytesIO(out_bytes), sheet_name="Census Input", header=6)
        st.subheader("Converted output preview (first 15 rows)")
        st.dataframe(out_preview.head(15), use_container_width=True)
    except Exception as pr_err:
        st.warning(f"Output preview warning: {pr_err}")

    st.download_button(
        label="Download converted file",
        data=out_bytes,
        file_name=out_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.markdown("---")
st.caption(
    "Caching enabled for preview/parse/normalize/convert. Filename uses C5 + today’s date. "
    "Output cells: A1/D1 Company, A2/D2 Address, A3/D3 FEIN, A4/D4 SIC, A5 totals. "
    "Headers on row 7; data starts row 8."
)
